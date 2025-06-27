import os
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import messagebox, filedialog
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
from functools import wraps
import logging
import sys
from typing import Optional,Tuple, Dict, List, Union, Any
from contextlib import contextmanager
from pathlib import Path

# Constants for cleaning and processing data

# Characters to remove from text to clean it up (e.g., extra spaces, quotes, or special characters).
NORMALIZE_CHAR = [
    "_x000D_",  # Excel carriage return
    "\r",       # Carriage return
    "\n",       # New line
    "\t",       # Tab
    "\f",       # Form feed
    "\v",       # Vertical tab
    '"',        # Double quote
    "＊",       # Japanese asterisk
    "'"         # Single quote
]

# Values to standardize or ignore during data processing (e.g., empty or zero values).
NORMALIZE_PATTERN = [
    None,           # Missing value
    "0",            # Zero
    "0:00",         # Zero time (short)
    "00:00:00",     # Zero time (full)
    "12:00:00午前"  # Japanese midnight (AM)

]

# Date formats for parsing dates and times, including Japanese AM/PM styles.
DATE_PATTERN = [
    '%Y/%m/%d %H:%M:%S',       # YYYY/MM/DD HH:MM:SS
    '%Y/%m/%d %H:%M:%S 午前',  # YYYY/MM/DD HH:MM:SS AM (Japanese)
    '%Y/%m/%d %H:%M:%S 午後',  # YYYY/MM/DD HH:MM:SS PM (Japanese)
    '%Y/%m/%d %H:%M',          # YYYY/MM/DD HH:MM
    '%Y/%m/%d %H:%M 午前',     # YYYY/MM/DD HH:MM AM (Japanese)
    '%Y/%m/%d %H:%M 午後',     # YYYY/MM/DD HH:MM PM (Japanese)
    '%Y/%m/%d',                 # YYYY/MM/DD
    '%Y-%m-%d'                  # YYYY-MM-DD
]

# Pairs of values to treat as the same or ignore during comparison, often for employment status or shifts.
IGNORED_PAIRS = [
    ("休み", "シフト時間コード-1"),                
    ("None", "シフト時間コード-1"),              
    (None, "シフト時間コード-1"),    
    ("", "シフト時間コード-1"),               
    ("フリー", "シフト時間コード2147483647"),   
    ("退職", "【退職後】"),                      
    ("退職", "退職後"),                         
    ("フリー", "シフト時間コード2147483647"),     
    ("ステム未使用期間", "【採用前】"),           
    ("ステム未使用期間", "採用前"),               
    ("システム未使用期間", "【採用前】"),         
    ("システム未使用期間", "採用前"),             
    ("長期休暇：傷病", "【長期休暇】傷病"),       
    ("長期休暇：", "【長期休暇】"),               
    ("【長期休暇】育児", "長期休暇：育児"),       
    ("採用前", "【採用前】"),                     
    ("【長期休暇】産後", "長期休暇：産後"),       
    ("【長期休暇】産前", "長期休暇：産前"),       
    ("システム未使用期間", ""),                   
    ("システム未使用期間", "None"),               
    ("システム未使用期間", None),                 
    ("【勤務時間帯項目合計】", None),  
    ("【勤務時間帯項目合計】", ""),           
    ("【勤務時間帯項目合計】", "None"),           
    ("退職", "【退職後】【長期休暇】育児")       
]

FINAL_NORMALIZE_CHAR = {
    "〜": "~",
    "～": "~",
    "：": ":",
    " ": "",
    '　': "",
}

# --- Constants for Header Labels ---



HEADER_SHIFT_NAME = "シフト - シフト" # V2 => D
HEADER_SHIFT_TIME_RANGE = "シフト - 時間帯" # V2 => E
HEADER_LEAVE_TYPE = "出欠 - 休暇種別・区分" # v2 => G
HEADER_PAID_LEAVE_DAYS = "出欠 - 有給\n日数" # V2 => H
HEADER_PAID_LEAVE_HOURS = "出欠 - 有給\n時間" # V2 => I
HEADER_TIMECARD_RANGE = "タイムカード - 時間帯" # V2 => L
HEADER_OUTING_TIME = "外出時間" # V2 => M
HEADER_TOTAL_WORK_TIME = "合計\n勤務\n時間" # V2 => N
HEADER_REGULAR_WORK_HOURS = "通常勤務 - 勤務\n時間" # V2 => O
HEADER_BREAK_TIME = "通常勤務 - 休憩\n時間" # V2 = P
HEADER_OFF_DUTY_TIME = "通常勤務 - 勤務外\n時間" #V2 => Q
HEADER_REGULAR_WORK_TIME = "通常勤務 - 時間帯" # V2 => R
HEADER_OVERTIME_HOURS = "時間外勤務 - 勤務\n時間" # V2 => S
HEADER_OVERTIME_WORK_HOURS = "時間外勤務 - 時間帯" # V2 => T

HEADER_LEAVE_APPLICATION = "申請書 - 休暇" # V2 = AD
HEADER_REMARK = "その他 - 備考" # V2 = AG



def setup_logging(debug_level: str = 'INFO', logger_name: Optional[str] = None) -> logging.Logger:
    """
    Sets up logging with both file and stream handlers.

    Args:
        debug_level (str): Logging level ('DEBUG', 'INFO', 'WARNING', etc.).
        logger_name (Optional[str]): Name for the logger. If None, root logger is used.

    Returns:
        logging.Logger: Configured logger object.
    """
    # Prepare log directory
    logs_dir = Path('logs')
    logs_dir.mkdir(parents=True, exist_ok=True)

    # Prepare timestamped log file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = logs_dir / f'excel_comparison_{timestamp}.log'

    # Define log level safely
    level = getattr(logging, debug_level.upper(), logging.INFO)

    # Create logger
    logger = logging.getLogger(logger_name)
    logger.setLevel(level)
    logger.handlers.clear()  # Avoid duplicate handlers if re-run

    # Define formatter
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

    # File handler
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    # Stream (console) handler
    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    logger.info(f'Logging initialized at level: {debug_level.upper()}')

    return logger


@contextmanager
def create_root():
    """
    Create and manage a Tkinter root window.
    
    Yields:
        Tk: The Tkinter root window.
    """
    root = tk.Tk()
    root.withdraw()
    try:
        yield root
    finally:
        root.destroy()
    

def select_directory(prompt: str) -> str:
    """
    Display a directory selection dialog and return the selected folder path.

    Args:
        prompt (str): The message to display in the dialog box.

    Returns:
        str: Path of the selected directory, or empty string if no directory is selected or an error occurs.
    """
    # Validate prompt input
    if not prompt or not isinstance(prompt, str):
        return ""

    try:
        # Show info dialog with the prompt
        messagebox.showinfo('情報', prompt)
        
        # Open directory selection dialog
        folder_selected = filedialog.askdirectory(title=prompt)
        
        # Check if a directory was selected
        if not folder_selected:
            return ""
        
        # Verify the selected path exists and is a directory
        if not os.path.isdir(folder_selected):
            return ""
        
        return folder_selected

    except Exception as e:
        logging.error(f"Error during directory selection: {str(e)}")
        return ""
    

def show_message(title: str, message: str) -> bool:
    """
    Display a simple message box with a title and message.

    Args:
        title (str): The title of the message box.
        message (str): The message to display in the box.

    Returns:
        bool: True if the message box is shown successfully, False if an error occurs.
    """
    # Check if title and message are valid strings
    if not title or not isinstance(title, str):
        logging.error(f"Invalid or empty title provided for message box: {title}")
        return False
    if not message or not isinstance(message, str):
        logging.error(f"Invalid or empty message provided for message box: {message}")
        return False

    try:
        # Show the message box
        messagebox.showinfo(title, message)
        return True
    except Exception as e:
        logging.error(f"Error displaying message box: {str(e)}")
        return False

    
def extract_sheet_name_string(sheet_name: str) -> Optional[str]:
    """
    Extract the meaningful part of a sheet name by removing leading numbers, dots, underscores, or hyphens.

    Args:
        sheet_name (str): The sheet name to process.

    Returns:
        str: The cleaned sheet name, or None if an error occurs or input is invalid.
    """
    # Check if sheet_name is a valid string
    if not sheet_name or not isinstance(sheet_name, str):
        # logging.error("Invalid or empty sheet name provided")
        return None

    try:
        # Remove leading numbers, dots, underscores, or hyphens
        cleaned_name = re.sub(r'^[\d._\-]+', '', sheet_name)
        
        # Check if the result is empty after cleaning
        if not cleaned_name:
            # logging.warning(f"No meaningful sheet name after cleaning: {sheet_name}")
            return None
        
        return cleaned_name

    except Exception as e:
        logging.error(f"Error processing sheet name '{sheet_name}': {str(e)}")
        return None
    

def get_col_header_pairs(arr1: np.ndarray, arr2: np.ndarray) -> Tuple[List[Tuple[int, int]], Dict[int, str], Dict[int, str]]:
    """
    Compare column headers from two arrays and return matching header pairs and header dictionaries.

    Args:
        arr1 (np.ndarray): First input array (e.g., from an Excel sheet).
        arr2 (np.ndarray): Second input array (e.g., from another Excel sheet).

    Returns:
        Tuple[List[Tuple[int, int]], Dict[int, str], Dict[int, str]]:
            - List of (col1, col2) pairs where headers match.
            - Dictionary mapping column indices to headers for arr1.
            - Dictionary mapping column indices to headers for arr2.
    """
    # Check if inputs are valid NumPy arrays with enough rows
    if not isinstance(arr1, np.ndarray) or not isinstance(arr2, np.ndarray):
        # logging.error("Inputs must be NumPy arrays")
        return [], {}, {}
    if arr1.shape[0] < 9 or arr2.shape[0] < 9:
        # logging.error("Arrays must have at least 9 rows for header extraction")
        return [], {}, {}

    try:
        # Initialize dictionaries to store headers
        header1 = {}
        header2 = {}
        current_main_header_v1 = None
        current_main_header_v2 = None

        # Get the maximum number of columns to process
        col_max = max(arr1.shape[1], arr2.shape[1])

        # Loop through columns starting from index 3
        for col in range(3, col_max):
            # Get main and sub headers from row 7 and 8, or None if out of bounds
            main1 = arr1[7, col] if col < arr1.shape[1] and arr1.shape[0] > 7 else None
            sub1 = arr1[8, col] if col < arr1.shape[1] and arr1.shape[0] > 8 else None
            main2 = arr2[7, col] if col < arr2.shape[1] and arr2.shape[0] > 7 else None
            sub2 = arr2[8, col] if col < arr2.shape[1] and arr2.shape[0] > 8 else None

            # Standardize specific header text (e.g., for overtime headers)
            main1 = "時間外割増振分" if main1 == "時間外割増振分け" else main1
            main2 = "時間外割増振分" if main2 == "時間外割増振分け" else main2

            # Update current main headers if new ones are found
            current_main_header_v1 = main1 if main1 else current_main_header_v1
            current_main_header_v2 = main2 if main2 else current_main_header_v2

            # Combine main and sub headers with a hyphen if both exist
            combined1 = f"{current_main_header_v1} - {sub1}" if sub1 and current_main_header_v1 else current_main_header_v1
            combined2 = f"{current_main_header_v2} - {sub2}" if sub2 and current_main_header_v2 else current_main_header_v2

            # Store unique headers in dictionaries
            if combined1 and combined1 not in header1.values():
                header1[col] = combined1
            if combined2 and combined2 not in header2.values():
                header2[col] = combined2

        # Find matching header pairs across the two arrays
        header_pairs = [
            (col1, col2)
            for col1, h1 in header1.items()
            for col2, h2 in header2.items()
            if h1 == h2
        ]

        return header_pairs, header1, header2

    except Exception as e:
        logging.error(f"Error processing header pairs: {str(e)}")
        return [], {}, {}
    
    

def get_row_header_pairs(arr1: np.ndarray, arr2: np.ndarray) -> Tuple[List[Tuple[int, int]], Dict[int, str], Dict[int, str]]:
    """
    Compare row headers from two arrays and return matching header pairs and header dictionaries.

    Args:
        arr1 (np.ndarray): First input array (e.g., from an Excel sheet).
        arr2 (np.ndarray): Second input array (e.g., from another Excel sheet).

    Returns:
        Tuple[List[Tuple[int, int]], Dict[int, str], Dict[int, str]]:
            - List of (row1, row2) pairs where headers match.
            - Dictionary mapping row indices to headers for arr1.
            - Dictionary mapping row indices to headers for arr2.
    """
    # Check if inputs are valid NumPy arrays with enough columns
    if not isinstance(arr1, np.ndarray) or not isinstance(arr2, np.ndarray):
        # logging.error("Inputs must be NumPy arrays")
        return [], {}, {}
    if arr1.shape[1] < 3 or arr2.shape[1] < 3:
        # logging.error("Arrays must have at least 3 columns for header extraction")
        return [], {}, {}
    if arr1.shape[0] < 10 or arr2.shape[0] < 10:
        # logging.error("Arrays must have at least 10 rows for header extraction")
        return [], {}, {}

    try:
        # Initialize dictionaries to store headers
        headers1 = {}
        headers2 = {}
        stop1 = stop2 = False

        # Loop through rows starting from index 9
        for row in range(9, min(arr1.shape[0], arr2.shape[0])):
            # Get header values from column 2
            val1 = arr1[row, 2] if row < arr1.shape[0] and arr1.shape[1] > 2 else None
            val2 = arr2[row, 2] if row < arr2.shape[0] and arr2.shape[1] > 2 else None

            # Store non-empty headers unless stopped
            if val1 and not stop1:
                headers1[row] = str(val1)  # Convert to string for consistency
            if val2 and not stop2:
                headers2[row] = str(val2)  # Convert to string for consistency

            # Stop collecting headers when '計' (total) is found
            if val1 == '計':
                stop1 = True
            if val2 == '計':
                stop2 = True

            # Exit loop if both arrays reach '計'
            if stop1 and stop2:
                break

        # Find matching header pairs
        header_pairs = [
            (row1, row2)
            for row1, v1 in headers1.items()
            for row2, v2 in headers2.items()
            if v1 == v2
        ]

        return header_pairs, headers1, headers2

    except Exception as e:
        logging.error(f"Error processing row header pairs: {str(e)}")
        return [], {}, {}
    

def normalize_time_format(time_str: Optional[str | int | float]) -> str:
    """
    Convert a time string to a standard HH:MM format or return empty string if invalid.

    Args:
        time_str (str | int | float | None): The time value to normalize (e.g., "9:30", "09:30:00", 930).

    Returns:
        str: Normalized time in "HH:MM" format (e.g., "9:30"), or empty string if invalid.
    """
    # Check for None, empty, or specific invalid values
    if time_str is None or str(time_str).strip().lower() in ['none', '', '出勤']:
        # logging.debug(f"Skipped normalization for invalid input: {time_str}")
        return ""

    try:
        # Convert to string and clean up separators
        time_str = str(time_str).replace("：", ":").replace("〜", "~").replace("～", "~").strip()

        # Skip if contains range indicator
        if '~' in time_str:
            # logging.debug(f"Skipped normalization due to range indicator: {time_str}")
            return ""

        # Handle time formats with seconds (e.g., "09:30:00" -> "09:30")
        if time_str.count(':') == 2:
            time_str = ':'.join(time_str.split(':')[:2])

        # Process time with hours and minutes
        if ':' in time_str:
            hours, minutes = time_str.split(':')
            hours = str(int(hours.lstrip('0') or '0'))  # Remove leading zeros, default to 0
            minutes = minutes.zfill(2)  # Pad minutes with leading zero if needed

            normalized_time = f"{hours}:{minutes}"
            # logging.debug(f"Normalized time: {time_str} -> {normalized_time}")
            return normalized_time

        # logging.warning(f"Invalid time format, no colon found: {time_str}")
        return ""

    except Exception as e:
        logging.error(f"Error normalizing time '{time_str}': {str(e)}")
        return ""

def normalize_time_range_symbols(time_str: Optional[str | int | float]) -> Optional[str]:
    """
    Clean a time string by replacing Japanese range symbols (〜, ～) with a standard tilde (~) and stripping whitespace.

    Args:
        time_str (str | int | float | None): The time string to normalize (e.g., "9:00〜17:00").

    Returns:
        str: Normalized string with standard tilde, or None if input is invalid.
    """
    # Check for None, empty, or specific invalid values
    if time_str is None or str(time_str).strip().lower() in ['none', '', '出勤']:
        # logging.debug(f"Skipped normalization for invalid input: {time_str}")
        return None

    try:
        # Convert to string, replace Japanese range symbols, and strip whitespace
        normalized = str(time_str).replace('〜', '~').replace('～', '~').strip()
        # logging.debug(f"Normalized time range symbols: '{time_str}' -> '{normalized}'")
        return normalized

    except Exception as e:
        logging.error(f"Error normalizing time range symbols '{time_str}': {str(e)}")
        return None
    

def format_time_range(time_str: Optional[str | int | float]) -> str:
    """
    Format a time range string into a standard HH:MM~HH:MM format.

    Args:
        time_str (str | int | float | None): The time range to format (e.g., "9:00〜17:00", "9~17").

    Returns:
        str: Formatted time range (e.g., "9:00~17:00"), or empty string if invalid.
    """
    # Normalize time range symbols first
    time_str = normalize_time_range_symbols(time_str)
    if not time_str:
        # logging.debug(f"Skipped formatting due to invalid input: {time_str}")
        return ""

    try:
        # Match time range pattern (e.g., "9:00~17:00" or "9~17")
        match = re.match(r'^\s*(\d{1,2}(:\d{2})?)\s*~\s*(\d{1,2}(:\d{2})?)\s*$', time_str)
        if not match:
            # logging.warning(f"Invalid time range format: {time_str}")
            return ""

        # Extract start and end times
        start_raw, _, end_raw, _ = match.groups()

        # Normalize start and end times to HH:MM format
        start_time = normalize_time_format(start_raw if ':' in start_raw else f"{start_raw}:00")
        end_time = normalize_time_format(end_raw if ':' in end_raw else f"{end_raw}:00")

        # Check if both times are valid
        if start_time and end_time:
            formatted_time = f"{start_time}~{end_time}"
            # logging.debug(f"Formatted time range: '{time_str}' -> '{formatted_time}'")
            return formatted_time

        # logging.warning(f"Failed to normalize time range: {time_str}")
        return ""

    except Exception as e:
        logging.error(f"Error formatting time range '{time_str}': {str(e)}")
        return ""


def normalize_value(value: Optional[Union[str, int, float, datetime]]) -> Union[str, datetime, int]:
    """
    Normalize a value by cleaning strings, handling datetimes, and converting numbers.

    Args:
        value (str | int | float | datetime | None): The value to normalize (e.g., string, number, or datetime).

    Returns:
        str | datetime | int: Normalized value (datetime, string, or integer), or empty string if invalid.
    """
    # Return datetime objects unchanged
    if isinstance(value, datetime):
        # logging.debug(f"Returning unchanged datetime: {value}")
        return value.strftime('%Y/%m/%d %H:%M:%S')

    # Handle None, empty, or 'none' values
    if value is None or str(value).strip().lower() in ['', 'none']:
        # logging.debug(f"Returning empty string for invalid input: {value}")
        return ""

    try:
        # Convert to string for processing
        value = str(value)

        # Remove special characters from NORMALIZE_CHAR
        for char in NORMALIZE_CHAR:
            value = value.replace(char, '')

        # Check if value is empty after removing special characters
        if not value.strip():
            # logging.debug(f"Returning empty string after cleaning: {value}")
            return ""

        # Normalize whitespace and remove leading/trailing spaces
        value = ' '.join(value.split()).strip()

        # Remove space before opening parenthesis (e.g., "9:00 (" -> "9:00(")
        value = re.sub(r'(\d{1,3}:\d{2})\s+\(', r'\1(', value)

        # Check if value is empty after normalization
        if not value:
            # logging.debug(f"Returning empty string after normalization: {value}")
            return ""

        # Return empty string for values in NORMALIZE_PATTERN
        if value in NORMALIZE_PATTERN:
            # logging.debug(f"Returning empty string for pattern match: {value}")
            return ""

        # Try converting to a number (integer if possible)
        try:
            num = float(value)
            if num.is_integer():
                # logging.debug(f"Converted to integer: '{value}' -> {int(num)}")
                return str(int(num))
            # logging.debug(f"Keeping as string, not an integer: {value}")
            return value
        except ValueError:
            # logging.debug(f"Returning normalized string: {value}")
            return value

    except Exception as e:
        logging.error(f"Error normalizing value '{value}': {str(e)}")
        return ""
    

def is_datetime_string(value: Optional[str]) -> bool:
    """
    Check if a string can be parsed as a datetime using predefined date patterns.

    Args:
        value (str | None): The string to check for datetime format.

    Returns:
        bool: True if the string matches a datetime pattern, False otherwise.
    """
    # Check if value is a valid string
    if not isinstance(value, str) or not value.strip():
        # logging.debug(f"Input is not a valid string: {value}")
        return False

    try:
        # Try parsing the string with each pattern in DATE_PATTERN
        for pattern in DATE_PATTERN:
            try:
                datetime.strptime(value, pattern)
                # logging.debug(f"String '{value}' matches datetime pattern '{pattern}'")
                return True
            except ValueError:
                continue

        # logging.debug(f"String '{value}' does not match any datetime pattern")
        return False

    except Exception as e:
        logging.error(f"Error checking datetime string '{value}': {str(e)}")
        return False


def extract_date_part(value: Union[str, int, float]) -> Optional[str]:
    """
    Extracts and standardizes date from various input formats to 'YYYY/MM/DD' format.
    
    Args:
        value: Input value that may contain a date (string, int, or float).
        
    Returns:
        str: Formatted date string in 'YYYY/MM/DD' format, or original value if no valid date found.
        
    Examples:
        >>> extract_date_part('2025-06-05')
        '2025/06/05'
        >>> extract_date_part('05/06/2025')
        '2025/06/05'
        >>> extract_date_part('invalid')
        'invalid'
    """
    if value is None or not str(value).strip():
        return str(value)

    # Try predefined datetime patterns
    for pattern in DATE_PATTERN:
        try:
            dt = datetime.strptime(str(value).strip(), pattern)
            return dt.strftime('%Y/%m/%d')
        except (ValueError, TypeError):
            continue
    
    # Fallback to regex for flexible date matching
    date_pattern = r'(?P<year>\d{4})[-/](?P<month>\d{1,2})[-/](?P<day>\d{1,2})'
    match = re.search(date_pattern, str(value).strip())
    
    if match:
        try:
            year = int(match.group('year'))
            month = int(match.group('month'))
            day = int(match.group('day'))
            
            # Validate date ranges
            if 1 <= month <= 12 and 1 <= day <= 31 and 1000 <= year <= 9999:
                # Additional validation for days in month
                datetime(year, month, day)  # Will raise ValueError if invalid
                return f"{year}/{month:02d}/{day:02d}"
        except (ValueError, TypeError):
            pass
            
    return str(value)


def is_time_string(value: Union[str, int, float]) -> bool:
    """
    Validates if the input represents a valid 24-hour time format.
    
    Supports formats like: 'HH:MM', 'H:MM', 'HHMM', 'HMM', 'HH.MM', etc.
    
    Args:
        value: Input value to check (string, int, or float).
        
    Returns:
        bool: True if the input is a valid time string, False otherwise.
        
    Examples:
        >>> is_time_string('14:30')
        True
        >>> is_time_string('830')
        True
        >>> is_time_string('25:00')
        False
        >>> is_time_string('invalid')
        False
    """
    if value is None or value is False or not str(value).strip():
        return False
    
    value = normalize_time_format(value)
    if not value:
        return False
    
    # Regex for common time formats: HH:MM, H:MM, HHMM, HMM
    time_pattern = r'^(?P<hours>\d{1,2})[:]?(?P<minutes>\d{2})$'
    match = re.search(time_pattern, value, re.IGNORECASE)
    
    if match:
        try:
            hours = int(match.group('hours'))
            minutes = int(match.group('minutes'))
            return 0 <= hours <= 23 and 0 <= minutes <= 59
        except (ValueError, TypeError):
            return False
    
    # Fallback for simple digit check (3 or 4 digits)
    clean_value = value.replace(':', '').strip()
    if clean_value.isdigit() and 3 <= len(clean_value) <= 4:
        try:
            if len(clean_value) == 3:
                hours = int(clean_value[0])
                minutes = int(clean_value[1:])
            else:
                hours = int(clean_value[:2])
                minutes = int(clean_value[2:])
            return 0 <= hours <= 23 and 0 <= minutes <= 59
        except (ValueError, TypeError):
            return False
    
    return False

def compare_time_parts(time1: Union[str, int, float], time2: Union[str, int, float]) -> bool:
    """
    Compares two time values to determine if they represent the same time.
    
    Args:
        time1: First time value (string, int, or float).
        time2: Second time value (string, int, or float).
        
    Returns:
        bool: True if both times represent the same valid time, False otherwise.
        
    Examples:
        >>> compare_time_parts('14:30', '1430')
        True
        >>> compare_time_parts('8:30', '08:30')
        True
        >>> compare_time_parts('25:00', '14:30')
        False
        >>> compare_time_parts('invalid', '14:30')
        False
    """
    # Normalize both time inputs
    normalized_time1 = normalize_time_format(time1)
    normalized_time2 = normalize_time_format(time2)
    
    
    # Return False if either time is invalid
    if normalized_time1 is None or not normalized_time1 or normalized_time2 is None or not normalized_time2:
        return False
    
    # Compare normalized times
    return normalized_time1 == normalized_time2


def compare_time_values(time1: Union[str, int, float], time2: Union[str, int, float]) -> bool:
    """
    Compares two time values to determine if they represent the same time, accounting for different formats.
    Treats times with and without seconds as equal if hours and minutes match (e.g., '08:00:00' == '8:00').
    
    Args:
        time1: First time value (string, int, or float).
        time2: Second time value (string, int, or float).
        
    Returns:
        bool: True if both times represent the same time (ignoring seconds if not provided), False otherwise.
        
    Examples:
        >>> compare_time_values('08:00:00', '8:00')
        True
        >>> compare_time_values('14:30:45', '1430')
        True
        >>> compare_time_values('14:30:45', '14:30:00')
        True
        >>> compare_time_values('14:30', '14:31')
        False
        >>> compare_time_values('invalid', '14:30')
        False
    """
    # Normalize both time inputs
    normalized_time1 = normalize_time_format(time1)
    normalized_time2 = normalize_time_format(time2)
    
    # Check if both inputs are valid time strings
    if not (is_time_string(time1) and is_time_string(time2)):
        return False
    
    # Compare hours and minutes only (ignore seconds)
    time1_hm = normalized_time1[:5]  # Extract HH:MM
    time2_hm = normalized_time2[:5]  # Extract HH:MM
    return time1_hm == time2_hm


def is_ignored_mismatch(value1: Any, value2: Any, case_sensitive: bool = False) -> bool:
    """
    Checks if two values form a pair that should be ignored as a mismatch.
    
    Args:
        value1: First value to compare.
        value2: Second value to compare.
        case_sensitive: If True, performs case-sensitive comparison (default: False).
        
    Returns:
        bool: True if the values form an ignored pair, False otherwise.
        
    Examples:
        >>> is_ignored_mismatch("yes", "true")
        True
        >>> is_ignored_mismatch("YES", "true")
        True
        >>> is_ignored_mismatch("yes", "no")
        False
        >>> is_ignored_mismatch(None, "")
        True
    """
    # Handle None or empty inputs
    if value1 is None and value2 is None:
        return True
    if value1 is None or value2 is None:
        return (value1, value2) in IGNORED_PAIRS or (value2, value1) in IGNORED_PAIRS
    
    # Convert to strings for consistent comparison
    try:
        str_value1 = str(value1).strip()
        str_value2 = str(value2).strip()
    except (TypeError, ValueError):
        return False
    
    # Apply case sensitivity
    if not case_sensitive:
        str_value1 = str_value1.lower()
        str_value2 = str_value2.lower()
    
    # Check if the pair is in IGNORED_PAIRS (in either order)
    for pair in IGNORED_PAIRS:
        # Convert pair values to strings for comparison
        pair1, pair2 = (str(p).strip().lower() if not case_sensitive else str(p).strip() for p in pair)
        
        if ((str_value1 == pair1 and str_value2 == pair2) or 
            (str_value1 == pair2 and str_value2 == pair1)):
            return True
    
    return False

def get_col_index(key_list: list[str], value_list: list[str], column_name: str) -> str:
    try:
        return key_list[value_list.index(column_name)]
    except ValueError:
        raise ValueError(f"Column '{column_name}' not found in value_list.")
    except IndexError:
        raise IndexError(f"Index {index} is out of bounds for key_list of length {len(key_list)}.")


def get_row_index(key_list: list[str], value_list: list[str], column_name: str) -> str:
    """
    Safely get the corresponding key from key_list where column_name matches in value_list.

    Args:
        key_list (list[str]): List of keys (e.g., Excel column letters or indexes).
        value_list (list[str]): List of column headers.
        column_name (str): The column name to search for.

    Returns:
        str: The corresponding key from key_list.

    Raises:
        ValueError: If column_name is not found in value_list.
        IndexError: If the index is out of range in key_list.
    """
    try:
        index = value_list.index(column_name)
        return str(key_list[index])  # Explicit type casting
    except ValueError:
        raise ValueError(f"Column '{column_name}' not found in value_list.")
    except IndexError:
        raise IndexError(f"Index {index} is out of bounds for key_list of length {len(key_list)}.")


def is_valid_time_range(s: str) -> bool:
    """
    Checks if a string is a valid time range in the format 'HH:MM ~ HH:MM'.
    
    Args:
        s (str): String to check.
        
    Returns:
        bool: True if the string is a valid time range, False otherwise.
        
    Examples:
        >>> is_valid_time_range('09:00~17:00')
        True
        >>> is_valid_time_range('0900~1700')
        True
        >>> is_valid_time_range('9:00~17:00')
        True
        >>> is_valid_time_range('09:00~17:00:00')
        False
    """
    return isinstance(s, str) and re.match(r'^\d{1,2}[:：]?\d{2}\s*~\s*\d{1,2}[:：]?\d{2}$', s)

def times_overlap(time_range1: Union[str, None], time_range2: Union[str, None]) -> bool:
    """
    Checks if two time ranges overlap in a 24-hour format.
    
    Args:
        time_range1: First time range (e.g., '09:00~17:00', '0900~1700').
        time_range2: Second time range (e.g., '12:00~18:00', '1200~1800').
        
    Returns:
        bool: True if the time ranges overlap, False otherwise.
        
    Examples:
        >>> times_overlap('09:00~17:00', '12:00~18:00')
        True
        >>> times_overlap('09:00~12:00', '13:00~18:00')
        False
        >>> times_overlap('09:00:00~17:00:00', '9:00~17:00')
        True
        >>> times_overlap('invalid', '12:00~18:00')
        False
        >>> times_overlap(None, '12:00~18:00')
        False
    """
    # Normalize time ranges
    normalized_range1 = format_time_range(time_range1)
    normalized_range2 = format_time_range(time_range2)
    
    # Return False if either range is invalid
    if not normalized_range1 or not normalized_range2:
        # logging.debug(f"Invalid time ranges: {time_range1} -> {normalized_range1}, {time_range2} -> {normalized_range2}")
        return False
    
    try:
        # Parse start and end times
        start1_str, end1_str = normalized_range1.split('~')
        start2_str, end2_str = normalized_range2.split('~')
        
        start1 = datetime.strptime(start1_str, '%H:%M:%S')
        end1 = datetime.strptime(end1_str, '%H:%M:%S')
        start2 = datetime.strptime(start2_str, '%H:%M:%S')
        end2 = datetime.strptime(end2_str, '%H:%M:%S')
        
        # Ensure start time is before end time for each range
        if start1 >= end1 or start2 >= end2:
            # logging.debug(f"Invalid time range order: {normalized_range1}, {normalized_range2}")
            return False
        
        # Check for overlap: max(start1, start2) < min(end1, end2)
        overlap = max(start1, start2) < min(end1, end2)
        # logging.debug(f"Overlap check: {normalized_range1} vs {normalized_range2} -> {overlap}")
        return overlap
    
    except (ValueError, TypeError) as e:
        logging.debug(f"Error checking time overlap: {e}")
        return False


def final_normalize_char(value: Any) -> str:
    """
    Normalize a string by replacing specific characters based on FINAL_NORMALIZE_CHAR dictionary.
    
    Args:
        value: Input value to normalize (expected to be a string or convertible to string).
    
    Returns:
        str: Normalized string, or empty string if input is invalid.

    """
    
    if pd.isna(value) or not str(value).strip():
        return ""
    
    try:
        value = str(value)
        for char, replacement in FINAL_NORMALIZE_CHAR.items():
            value = value.replace(char, replacement)
        
        return value
    except (TypeError, ValueError) as e:
        # print(f"Failed to normalize value: {value}, error: {e}")
        logging.warning(f"Failed to normalize value: {value}, error: {e}")
        return ""
    
def normalize_string_date_time(value):
    """
    Normalize a time range string (e.g., '8:3~16:3') to a standard format (e.g., '8:03~16:03').
    
    Args:
        value: Input string representing a time range.
    
    Returns:
        str: Normalized time range (e.g., '8:03~16:03'), or empty string if invalid.
    """
    if pd.isna(value) or not str(value).strip() or str(value).strip().lower() in ['none', '', '出勤']:
        return ""

    try:
        # Standardize separators (handle ~, 〜, ～, and :)
        value = str(value).replace("〜", "~").replace("～", "~").replace("：", ":")
        # Remove spaces around separators
        value = re.sub(r'\s*[~:]\s*', lambda m: m.group(0).strip(), value).strip()

        # Replace ~ and : with commas for splitting
        normalized = re.sub(r'[~:]', ',', value)
        value_arr = [part.strip() for part in normalized.split(',') if part.strip()]

        if len(value_arr) != 4:  # Expecting [start_hour, start_min, end_hour, end_min]
            # logging.warning(f"Invalid time range format: {value}, expected 4 parts, got {value_arr}")
            return ""

        # Process each part
        processed_arr = []
        for i, val in enumerate(value_arr):
            if val.isdigit():
                # Remove leading zeros and convert to int
                num = int(val.lstrip("0") or "0")  # Handle '0' case
                # Ensure two-digit format for minutes (indices 1 and 3)
                if i in [1, 3]:
                    val = f"{num:02d}"
                else:
                    val = str(num)
            else:
                # logging.warning(f"Non-numeric part in time range: {val}")
                return ""
            processed_arr.append(val)

        # Reconstruct as HH:MM~HH:MM
        start_time = f"{processed_arr[0]}:{processed_arr[1]}"
        end_time = f"{processed_arr[2]}:{processed_arr[3]}"
        result = f"{start_time}~{end_time}"

        # Validate time format
        if not re.match(r'^\d{1,2}:\d{2}~\d{1,2}:\d{2}$', result):
            # logging.warning(f"Invalid normalized time format: {result}")
            return ""

        return result

    except (TypeError, ValueError) as e:
        logging.warning(f"Failed to normalize time string: {value}, error: {e}")
        return ""
    
def is_vacation_equivalent(value1: object, value2: object) -> bool:
    """
    Check whether two values are equivalent types of vacation labels.
    
    This function considers two values to be equivalent if both start with:
    - "【休暇" (with brackets), or
    - "休暇" (without brackets)

    Args:
        value1 (object): First value to compare (expected to be string-like).
        value2 (object): Second value to compare (expected to be string-like).

    Returns:
        bool: True if both values represent equivalent vacation types; False otherwise.
    """
    try:
        # Ensure inputs are strings for safe string operations
        str1 = str(value1).strip()
        str2 = str(value2).strip()

        # Compare with both bracketed and unbracketed formats
        return (
            str1.startswith("【休暇") and str2.startswith("【休暇")
        ) or (
            str1.startswith("休暇") and str2.startswith("休暇")
        )
    except Exception as e:
        logging.warning(f"[Error] Failed to compare vacation equivalence: {e}")
        return False
    
def flatten_and_normalize_lines(value: Optional[Union[str, List[str]]]) -> List[str]:
    """
    Flattens a multi-line string or list, splitting on line breaks and _x000D_, and normalizes each line.

    Args:
        value (str | List[str] | None): Input value to process.

    Returns:
        List[str]: A list of cleaned and normalized lines.
    """
    if value is None:
        return []

    # Convert to string if it's a list of strings or a single string
    if isinstance(value, list):
        value = '\n'.join(map(str, value))
    elif not isinstance(value, str):
        value = str(value)

    # Split on newlines and Excel-style `_x000D_` codes
    lines = re.split(r'(?:\r\n|\r|\n|_x000D_)+', value.strip())

    # Normalize and filter out empty lines
    return [normalize_value(line) for line in lines if line.strip()]

def check_multi_lines_value_equal(value1: Optional[Union[str, List[str]]],
                                  value2: Optional[Union[str, List[str]]]) -> bool:
    """
    Checks if two multi-line values are equal after normalization, ignoring order.

    Args:
        value1 (str | List[str] | None): First value.
        value2 (str | List[str] | None): Second value.

    Returns:
        bool: True if both are equal post-normalization; False otherwise.
    """
    if value1 is None or value2 is None:
        return False

    try:
        value1_list = flatten_and_normalize_lines(value1)
        value2_list = flatten_and_normalize_lines(value2)
    except Exception as e:
        logging.error("Error normalizing values: %s", e)
        return False

    # logging.debug("Normalized arrays:\n%s\nvs\n%s", value1_list, value2_list)

    return sorted(value1_list) == sorted(value2_list)


def parse_time_range(time_str: str) -> Optional[Tuple[int, int]]:
    """
    Parses a time range string in the format 'HH:MM 〜 HH:MM' and returns start and end time in minutes.
    """
    match = re.search(r'(\d{1,2}):(\d{2})\s*〜\s*(\d{1,2}):(\d{2})', time_str)
    if match:
        start_hour, start_minute, end_hour, end_minute = map(int, match.groups())
        start = start_hour * 60 + start_minute
        end = end_hour * 60 + end_minute
        return start, end
    return None

def check_overlapping(time1: str, time2: str) -> bool:
    """
    Returns True if time1 and time2 ranges overlap (partial overlap), otherwise False.
    Assumes time format is 'HH:MM 〜 HH:MM'.
    """
    time1 = normalize_value(time1)
    time2 = normalize_value(time2)

    range1 = parse_time_range(time1)
    range2 = parse_time_range(time2)

    if not range1 or not range2:
        return False  # invalid time formats

    start1, end1 = range1
    start2, end2 = range2

    # Overlap condition: partial or full
    return start1 < end2 and start2 < end1

    
def compare_excel_files(file1: Union[str, bytes], file2: Union[str, bytes], subfolderName: str) -> Tuple[Optional[str], Optional[openpyxl.Workbook], List[Dict]]:
    """
    Compares two Excel files and highlights mismatches in the second workbook.
    
    Args:
        file1: Path to the first Excel file (string or bytes).
        file2: Path to the second Excel file (string or bytes).
        subfolderName: Subfolder identifier for specific comparison rules.
        
    Returns:
        Tuple containing:
        - result: 'O' if no mismatches, 'X' if mismatches found, None if error.
        - wb2_openpyxl: Modified second workbook with highlighted mismatches, or None if error.
        - reports: List of mismatch reports per sheet, each containing sheet name, mismatch details, and count.
        
    Examples:
        >>> result, wb, reports = compare_excel_files('file1.xlsx', 'file2.xlsx', '50028001')
        >>> print(result)
        'X'  # If mismatches found
    """
    try:
        # Validate inputs
        if not all(isinstance(x, (str, bytes)) for x in [file1, file2]) or not isinstance(subfolderName, str):
            logging.error(f"Invalid input types: file1={type(file1)}, file2={type(file2)}, subfolderName={type(subfolderName)}")
            return None, None, []

        file1_actual_path = os.path.basename(file1)
        file2_actual_path = os.path.basename(file2)

        reports = []
        mismatch_found = 0
        fill_pattern_yellow = PatternFill(patternType="solid", fgColor='FFFF00')
        
        # Load workbooks
        wb1_openpyxl = openpyxl.load_workbook(file1, data_only=True)
        wb2_openpyxl = openpyxl.load_workbook(file2, data_only=True)
        
        # Map normalized sheet names to original names
        sheet_names1 = {extract_sheet_name_string(s): s for s in wb1_openpyxl.sheetnames}
        sheet_names2 = {extract_sheet_name_string(s): s for s in wb2_openpyxl.sheetnames}
        
        sheet_openpyxl = [(sheet.title, extract_sheet_name_string(sheet.title)) 
                         for sheet in wb2_openpyxl.worksheets 
                         if sheet.sheet_state == 'visible']
        sheet_openpyxl_dict = {string: orig for orig, string in sheet_openpyxl}
        
        common_sheets = set(sheet_names1) & set(sheet_names2)
        
        if not common_sheets:
            logging.warning("No common sheets found between files")
            return 'X', wb2_openpyxl, []
        
        for sheet_name in common_sheets:
            sheet_report = []
            sheet_mismatch_found = 0
            
            # Read sheets into DataFrames
            df1 = pd.read_excel(file1, sheet_name=sheet_names1[sheet_name], header=None, dtype=str).fillna("")
            df2 = pd.read_excel(file2, sheet_name=sheet_names2[sheet_name], header=None, dtype=str).fillna("")
            ws = wb2_openpyxl[sheet_openpyxl_dict[sheet_name]]
            
            arr1 = df1.to_numpy()
            arr2 = df2.to_numpy()
            
            # Get header pairs
            col_header_pairs, col_header1, col_header2 = get_col_header_pairs(arr1, arr2)
            row_header_pairs, row_header1, row_header2 = get_row_header_pairs(arr1, arr2)
            
            list_col_header1_value = list(col_header1.values())
            list_col_header2_value = list(col_header2.values())
            list_col_header1_key = list(col_header1.keys())
            list_col_header2_key = list(col_header2.keys())
            
            list_row_header1_value = list(row_header1.values())
            list_row_header2_value = list(row_header2.values())
            list_row_header1_key = list(row_header1.keys())
            list_row_header2_key = list(row_header2.keys())
            
            indexes_col = {
                "shift_name_col_v1": get_col_index(list_col_header1_key, list_col_header1_value, HEADER_SHIFT_NAME),   # D - V1
                "shift_name_col_v2": get_col_index(list_col_header2_key, list_col_header2_value, HEADER_SHIFT_NAME),   # D - V2
                "shift_time_range_col_v1": get_col_index(list_col_header1_key, list_col_header1_value, HEADER_SHIFT_TIME_RANGE),  # E - V1
                "leave_type_col_v1": get_col_index(list_col_header1_key, list_col_header1_value, HEADER_LEAVE_TYPE),   # G - V1
                "leave_type_col_v2": get_col_index(list_col_header2_key, list_col_header2_value, HEADER_LEAVE_TYPE),   # G - V2
                "paid_leave_days_col_v2": get_col_index(list_col_header2_key, list_col_header2_value, HEADER_PAID_LEAVE_DAYS),  # H - V2
                "paid_leave_hours_col_v2": get_col_index(list_col_header2_key, list_col_header2_value, HEADER_PAID_LEAVE_HOURS),  # I - V2
                "timecard_time_range_col_v2": get_col_index(list_col_header2_key, list_col_header2_value, HEADER_TIMECARD_RANGE),  # L - V2
                "outing_time_col_v2": get_col_index(list_col_header2_key, list_col_header2_value, HEADER_OUTING_TIME),  # M - V2
                "total_work_time_col_v2": get_col_index(list_col_header2_key, list_col_header2_value, HEADER_TOTAL_WORK_TIME),  # N - V2
                "regular_work_hours_col_v2" : get_col_index(list_col_header2_key, list_col_header2_value, HEADER_REGULAR_WORK_HOURS),  # O - V2
                "break_time_col_v1": get_col_index(list_col_header1_key, list_col_header1_value, HEADER_BREAK_TIME),   # P - V1
                "off_duty_time_col_v2": get_col_index(list_col_header2_key, list_col_header2_value, HEADER_OFF_DUTY_TIME),  # Q - V2
                'regular_work_time_col_v1': get_col_index(list_col_header1_key, list_col_header1_value, HEADER_REGULAR_WORK_TIME), # Q - V1
                'regular_work_time_col_v2': get_col_index(list_col_header2_key, list_col_header2_value, HEADER_REGULAR_WORK_TIME), # R - V2
                "overtime_hours_col_v2": get_col_index(list_col_header2_key, list_col_header2_value, HEADER_OVERTIME_HOURS),  # S - V2
                "overtime_work_hours_col_v2": get_col_index(list_col_header2_key, list_col_header2_value, HEADER_OVERTIME_WORK_HOURS),  # T - V2
                "leave_application_col_v1": get_col_index(list_col_header1_key, list_col_header1_value, HEADER_LEAVE_APPLICATION),  # Z - V1
                "remark_col_v1" : get_col_index(list_col_header1_key, list_col_header1_value, HEADER_REMARK),  # AF - V1
                "remark_col_v2" : get_col_index(list_col_header2_key, list_col_header2_value, HEADER_REMARK),  # AG - V2
            }
            
            excel_col_v1= {
                'D':'shift_name_col_v1',
                'E':'shift_time_range_col_v1',
                'G':'leave_type_col_v1',
                'P':'break_time_col_v1',
                'Q':'regular_work_time_col_v1',
                'Z':'leave_application_col_v1',
                'AF':"remark_col_v1"
            }

            excel_col_v2= {
                'D':'shift_name_col_v2',
                'G':'leave_type_col_v2',
                'H':'paid_leave_days_col_v2',
                'I':'paid_leave_hours_col_v2',
                'L':'timecard_time_range_col_v2',
                'M':'outing_time_col_v2',
                'N':'total_work_time_col_v2',
                'O':'regular_work_hours_col_v2',
                'Q':'off_duty_time_col_v2',
                'R':'regular_work_time_col_v2',
                'S':'overtime_hours_col_v2',
                'T':'overtime_work_hours_col_v2',
                "AG":"remark_col_v2"
            }
            
            for row1, row2 in row_header_pairs:
                
                # skip program issues 
                plan_row = list_row_header1_key[list_row_header1_value.index("計")]
                # break_time_col_v1 = indexes_col["break_time_col_v1"]
                
                
                
                if row1 == plan_row :
                    continue
                
                
                # Skip rows with multiple '時間休' in leave type
                try:
                    leave_type_col = indexes_col[excel_col_v1['G']]
                    leave_type_value = arr1[row1, leave_type_col]
                    
                    if leave_type_value and leave_type_value.count("時間休") >= 2:
                        continue
                    
                    # leave_time_col = indexes_col[excel_col_v2['G']]
                    # leave_time_value = arr2[row2, leave_time_col] 
                    
                    # outing_time_col = indexes_col[excel_col_v2['M']]
                    # outing_time_value = arr2[row2, outing_time_col]
                    
                    # if leave_time_value and outing_time_value:
                    #     leave_time = format_time_range(normalize_value(leave_time_value))
                    #     outing_time = format_time_range(normalize_value(outing_time_value))
                        
                    #     if is_valid_time_range(leave_time) and is_valid_time_range(outing_time):
                    #         if times_overlap(leave_time, outing_time):
                    #             continue
                    
                except (ValueError, IndexError) as e:
                    logging.warning(f"Could not find '出欠 - 休暇種別・区分' column in sheet {sheet_name}: {e}")
                    continue
                
                # Skip specific shift codes based on subfolderName
                try:
                    shift_value_v1_col = indexes_col["shift_name_col_v1"]
                    shift_value_v1 = arr1[row1, shift_value_v1_col]
                    
                    skip_conditions = {
                        '50028001': lambda x: x.startswith("シフト時間コード14"),
                        '50001008': lambda x: x.startswith("シフト時間コード1076"),
                        '50033001': lambda x: x in [
                            "シフト時間コード50", "シフト時間コード51", "シフト時間コード52",
                            "シフト時間コード53", "シフト時間コード54", "シフト時間コード55"
                        ],
                        '50071001': lambda x: x.startswith("シフト時間コード87")
                    }
                    if subfolderName in skip_conditions and skip_conditions[subfolderName](shift_value_v1):
                        continue
                except (ValueError, IndexError) as e:
                    logging.warning(f"Could not find 'シフト - シフト' column in sheet {sheet_name}: {e}")
                    continue
                
                try: 
                    shift_time_range_col_v1 = indexes_col["shift_time_range_col_v1"]
                    shift_time_range_value_v1 = normalize_value(arr1[row1, shift_time_range_col_v1])
                    
                    if shift_time_range_value_v1 in ["00~00", "00:00~00:00", "0:00～0:00", "00:00～00:00"]:
                        continue
                except (ValueError, IndexError) as e:
                    logging.warning(f"Could not find 'シフト - シフト' column in sheet {sheet_name}: {e}")
                    continue 
                
                
                try:
                    # get remark from v1
                    remark_col_v1 = indexes_col["remark_col_v1"]
                    remark_value_v1 = arr1[row1,remark_col_v1]
                    
                    #get remark from v2
                    remark_col_v2 = indexes_col["remark_col_v2"]
                    remark_value_v2 = arr2[row2,remark_col_v2]
                    
                    # check remark
                    if remark_value_v1.strip() == "システム未使用期間" and remark_value_v2.strip() in ["【採用前】",""]:
                        continue
                    
                except (ValueError, IndexError) as e:
                    logging.warning(f"Could not find 'その他 - 備考' column in sheet {sheet_name}: {e}")
                    continue 
                    
                
                
                
                for col1, col2 in col_header_pairs:
                    
                    # # skip program issues 
                    # plan_row = list_row_header1_key[list_row_header1_value.index("計")]
                    # break_time_col_v1 = indexes_col["break_time_col_v1"]
                    
                    
                    
                    # if row1 == plan_row and col1 == break_time_col_v1:
                    #     continue
                    
                    
                    value1 = arr1[row1, col1]
                    value2 = arr2[row2, col2]
                    
                    if pd.isna(value1) and pd.isna(value2):
                        continue
                    
                    if is_ignored_mismatch(value1,value2):
                        continue
                    
                    if not value1 and not value2:
                        continue
                    
                    if value1 == '' and value2 == '':
                        continue
                    
                    
                    # Handle multi-line values
                    if (("\n" in value1 or "_x000D_" in value1) and ("\n" in value2 or "_x000D_" in value2)) and check_multi_lines_value_equal(value1, value2):
                        logging.debug(f"Multi-line match at ({row2+1}, {col2+1}): {value1} vs {value2}")
                        continue
                           
                    value1 = normalize_value(value1)
                    value2 = normalize_value(value2)
                    

                    
                    
                    if not value1 and not value2:
                        continue
                    
                    logging.debug(f"{subfolderName}/{file1_actual_path}/{sheet_name} : ({row1+1},{col1+1}) : ({row2+1},{col2+1}) : Comparing value: {value1} vs {value2}")
                    
                    # if col is N or O 
                    if col2  == indexes_col[excel_col_v2['N']] or col2 == indexes_col[excel_col_v2['O']]:
                        
                        # leave time
                        g_col_v2 = indexes_col[excel_col_v2['G']]
                        g_col_value_v2 = arr2[row2, g_col_v2]
                        
                        # outing time
                        m_col_v2 = indexes_col[excel_col_v2['M']]
                        m_col_value_v2 = arr2[row2, m_col_v2]
                        
                        #check leave time and outing time is overlapping
                        if g_col_value_v2 and m_col_value_v2 :
                            if check_overlapping(g_col_value_v2, m_col_value_v2):
                                continue
                        
                        # overtime hours
                        s_col_v2 = indexes_col[excel_col_v2['S']]
                        s_col_value_v2 = arr2[row2, s_col_v2]
                         
                        # overtime work range
                        t_col_v2 = indexes_col[excel_col_v2['T']]
                        t_col_value_v2 = arr2[row2, t_col_v2]
                        
                        # check outing time and overtime work range or overtime hours is not empty
                        if m_col_value_v2 and (t_col_value_v2 or s_col_value_v2):
                            continue 
                    
                    if col2 == indexes_col[excel_col_v2['Q']]:
                        d_col = indexes_col[excel_col_v2['D']]
                        d_col_value = normalize_value(arr2[row2, d_col])
                        h_col = indexes_col[excel_col_v2['H']]
                        h_col_value = normalize_value(arr2[row2, h_col])
                        i_col = indexes_col[excel_col_v2['I']]
                        i_col_value = normalize_value(arr2[row2, i_col])
                        
                        if not d_col_value and h_col_value and i_col_value:
                            continue
                        
                        l_col = indexes_col[excel_col_v2['L']]
                        l_col_value = normalize_value(arr2[row2, l_col])
                        
                        if not l_col_value and  (len(l_col_value.split("～")) < 2) or (len(l_col_value.split("~")) < 2):
                            continue
                        
                        
                            
                    # Handle specific columns
                    try:
                        off_duty_time_col_v2 = indexes_col["off_duty_time_col_v2"]
                        overtime_hours_col_v2 = indexes_col["overtime_hours_col_v2"]
                    except (ValueError, IndexError) as e:
                        logging.warning(f"Could not find required columns in sheet {sheet_name}: {e}")
                        continue
                    
                    if col2 == off_duty_time_col_v2:
                        regular_work_off_duty_hours_v1 = normalize_time_format(value1)
                        regular_work_off_duty_hours_v2 = normalize_time_format(value2)
                        overtime_work_hours_v2 = normalize_time_format(normalize_value(arr2[row2, overtime_hours_col_v2]))
                        r_col_val = normalize_time_format(normalize_value(arr2[row2, overtime_hours_col_v2-1]))
                        
                        if regular_work_off_duty_hours_v1 is None and regular_work_off_duty_hours_v2 is not None:
                            if r_col_val is not None or overtime_work_hours_v2 is not None:
                                ws.cell(row=row2+1, column=col2+1).fill = fill_pattern_yellow
                                mismatch_found += 1
                                sheet_mismatch_found += 1
                                sheet_report.append({
                                    "row1": row1+1, "col1": col1+1, "val1": value1,
                                    "row2": row2+1, "col2": col2+1, "val2": value2
                                })
                                logging.info(f"Mismatch found at ({row2+1}, {col2+1}): {value1} vs {value2}")
                                continue
                        
                        try:
                            shift_col_v2 = indexes_col["shift_name_col_v2"]
                            paid_leave_days_col_v2 = indexes_col["paid_leave_days_col_v2"]
                            paid_leave_hours_col_v2 = indexes_col["paid_leave_hours_col_v2"]
                            shift_value_v2 = normalize_value(arr2[row2, shift_col_v2])
                            paid_leave_days_value_v2 = normalize_value(arr2[row2, paid_leave_days_col_v2])
                            paid_leave_hours_value_v2 = normalize_value(arr2[row2, paid_leave_hours_col_v2])
                            
                            if shift_value_v2 is None and paid_leave_days_value_v2 and paid_leave_hours_value_v2:
                                continue
                        except (ValueError, IndexError) as e:
                            logging.warning(f"Could not find shift or attendance columns in sheet {sheet_name}: {e}")
                            continue
                        
                        try:
                            timecard_time_period_col_v2 = indexes_col["timecard_time_range_col_v2"]
                            timecard_time_period_value_v2 = normalize_value(arr2[row2, timecard_time_period_col_v2])
                            if timecard_time_period_value_v2:
                                formatted_time = format_time_range(timecard_time_period_value_v2)
                                if formatted_time is None:
                                    logging.warning(f"Skipping invalid timecard time period: {timecard_time_period_value_v2}")
                                    continue
                        except (ValueError, IndexError) as e:
                            logging.warning(f"Could not find 'タイムカード - 時間帯' column in sheet {sheet_name}: {e}")
                            continue
                        
                        
                        
                    q_col_v1 = indexes_col[excel_col_v1['Q']]
                    s_col_v2 = indexes_col[excel_col_v2['S']]
                    r_col_v2 = s_col_v2 - 1
                    # Check the specific condition for V2勤務外時間 (Q列), V1勤務外時間 (Q列), and 時間外勤務.勤務時間 (S列)
                    if col1 == q_col_v1:  # Q列 (column 17)
                        v2_out_time = normalize_time_format(value1)
                        v1_out_time = normalize_time_format(value2)
                        r_col_val = normalize_time_format(arr2[row2, r_col_v2])
                        overtime_hours_v2 = normalize_time_format(normalize_value(arr2[row2, s_col_v2]))  # S列 (column 19)
                        
                        if (
                            (v2_out_time != "00:00" or v2_out_time.strip() != "" or v2_out_time is not None or v2_out_time != "None") and
                            (v1_out_time == "00:00" or v1_out_time.strip() == "" or v1_out_time is None or v1_out_time == "None") and
                            (r_col_val != "00:00" or r_col_val.strip() != "" or r_col_val is not None or r_col_val != "None")
                        ):
                            # logging.debug(
                            #     f"Skipping comparison for row {row2} due to specified conditions: "
                            #     f"V2勤務外時間={v2_out_time}, V1勤務外時間={v1_out_time}, 時間外勤務.勤務時間={r_col_val}"
                            # )
                            continue
                        
                        
                        if (
                            (v2_out_time != "00:00" or v2_out_time.strip() != "" or v2_out_time is not None or v2_out_time != "None") and
                            (v1_out_time == "00:00" or v1_out_time.strip() == "" or v1_out_time is None or v1_out_time == "None") and
                            (overtime_hours_v2 != "00:00" or overtime_hours_v2.strip() != "" or overtime_hours_v2 is not None or overtime_hours_v2 != "None")
                        ):
                            ws.cell(row=row2+1, column=col2+1).fill = fill_pattern_yellow
                            mismatch_found += 1
                            sheet_mismatch_found += 1
                            sheet_report.append({
                                "row1": row1+1, "col1": col1+1, "val1": value1,
                                "row2": row2+1, "col2": col2+1, "val2": value2
                            })
                            logging.info(f"Time overlap detected at row {row1+1}: 有給(時間休)={leave_time}, 外出={outing_time}")
                            continue
                        
                    
                    # Handle leave type column
                    if value1 is None and col1 == indexes_col['leave_type_col_v1']:
                        try:
                            if value2 == "その他(一日)":
                                if arr1[row1, col1-1] == "休み" and arr2[row2, col2-1] == "休み":
                                    continue
                            
                            leave_application_form_col_v1 = indexes_col['leave_application_col_v1']
                            if arr1[row1, leave_application_form_col_v1] == "✔":
                                continue
                        except (ValueError, IndexError) as e:
                            logging.warning(f"Could not find '申請書 - 休暇' column in sheet {sheet_name}: {e}")
                            continue
                    
                    # Handle shift time range
                    if col1 == indexes_col['shift_time_range_col_v1']:
                        leave_time = format_time_range(value1)
                        outing_time = format_time_range(normalize_value(arr1[row1, col1+1]))
                        
                        if is_valid_time_range(leave_time) and is_valid_time_range(outing_time):
                            if times_overlap(leave_time, outing_time):
                                ws.cell(row=row2+1, column=col2+1).fill = fill_pattern_yellow
                                mismatch_found += 1
                                sheet_mismatch_found += 1
                                sheet_report.append({
                                    "row1": row1+1, "col1": col1+1, "val1": value1,
                                    "row2": row2+1, "col2": col2+1, "val2": value2
                                })
                                logging.info(f"Time overlap detected at row {row1+1}: 有給(時間休)={leave_time}, 外出={outing_time}")
                                continue
                        # elif not (leave_time and outing_time):
                            # logging.warning(f"Skipping time overlap check due to invalid times: leave_time={value1}, outing_time={arr1[row1, col1+1]}")
                    
                    # Handle datetime values
                    is_datetime1 = isinstance(value1, datetime) or is_datetime_string(value1)
                    is_datetime2 = isinstance(value2, datetime) or is_datetime_string(value2)
                    
                    if is_datetime1 or is_datetime2:
                        date1 = extract_date_part(value1)
                        date2 = extract_date_part(value2)
                        
                        if date1 != date2:
                            ws.cell(row=row2+1, column=col2+1).fill = fill_pattern_yellow
                            mismatch_found += 1
                            sheet_mismatch_found += 1
                            sheet_report.append({
                                "row1": row1+1, "col1": col1+1, "val1": value1,
                                "row2": row2+1, "col2": col2+1, "val2": value2
                            })
                            logging.info(f"Date mismatch at ({row2+1}, {col2+1}): {value1} vs {value2}")
                        continue
                    
                    # Handle time values
                    is_time1 = is_time_string(str(value1))
                    is_time2 = is_time_string(str(value2))
                    
                    if is_time1 and is_time2:
                        if not compare_time_values(value1, value2):
                            ws.cell(row=row2+1, column=col2+1).fill = fill_pattern_yellow
                            mismatch_found += 1
                            sheet_mismatch_found += 1
                            sheet_report.append({
                                "row1": row1+1, "col1": col1+1, "val1": value1,
                                "row2": row2+1, "col2": col2+1, "val2": value2
                            })
                            logging.info(f"Time mismatch at ({row2+1}, {col2+1}): {value1} vs {value2}")
                        continue
                    elif is_time1 != is_time2:
                        ws.cell(row=row2+1, column=col2+1).fill = fill_pattern_yellow
                        mismatch_found += 1
                        sheet_mismatch_found += 1
                        sheet_report.append({
                            "row1": row1+1, "col1": col1+1, "val1": value1,
                            "row2": row2+1, "col2": col2+1, "val2": value2
                        })
                        logging.info(f"Time mismatch at ({row2+1}, {col2+1}): {value1} vs {value2}")
                        continue
                    
                    # Handle vacation equivalence
                    if is_vacation_equivalent(value1, value2):
                        logging.debug(f"Vacation-equivalent match at ({row2+1}, {col2+1}): {value1} vs {value2}")
                        continue
                    
                    # Handle time ranges
                    if any(separator in str(value1) or separator in str(value2) 
                           for separator in ['〜', '～', '~']):
                        time1_formatted = format_time_range(str(value1))
                        time2_formatted = format_time_range(str(value2))
                        
                        if time1_formatted and time2_formatted:
                            time1_parts = time1_formatted.split('~')
                            time2_parts = time2_formatted.split('~')
                            
                            if len(time1_parts) == 2 and len(time2_parts) == 2:
                                start_match = compare_time_parts(time1_parts[0], time2_parts[0])
                                end_match = compare_time_parts(time1_parts[1], time2_parts[1])
                                
                                if start_match and end_match:
                                    # logging.debug(f"Time range match at ({row2+1}, {col2+1}): {value1} vs {value2}")
                                    continue
                            
                            ws.cell(row=row2+1, column=col2+1).fill = fill_pattern_yellow
                            mismatch_found += 1
                            sheet_mismatch_found += 1
                            sheet_report.append({
                                "row1": row1+1, "col1": col1+1, "val1": value1,
                                "row2": row2+1, "col2": col2+1, "val2": value2
                            })
                            logging.info(f"Time range mismatch at ({row2+1}, {col2+1}): {value1} vs {value2}")
                            continue
                            
                    # Handle general value comparison
                    if str(value1) != str(value2):
                        if not is_ignored_mismatch(value1, value2):
                            if "【" in str(value1) and "】" in str(value1) and "【" in str(value2) and "】" in str(value2):
                                value1_clean = re.sub(r'【(.*?)】', lambda m: f"【{re.sub(r'\d+', '', m.group(1))}】", str(value1))
                                value2_clean = re.sub(r'【(.*?)】', lambda m: f"【{re.sub(r'\d+', '', m.group(1))}】", str(value2))
                                if value1_clean == value2_clean:
                                    logging.debug(f"Bracketed value match at ({row2+1}, {col2+1}): {value1} vs {value2}")
                                    continue
                            
                            value1_norm = final_normalize_char(value1)
                            value2_norm = final_normalize_char(value2)
                            
                            if value1_norm == value2_norm:
                                # logging.debug(f"Normalized char match at ({row2+1}, {col2+1}): {value1} vs {value2}")
                                continue
                            
                            value1_dt = normalize_string_date_time(value1)
                            value2_dt = normalize_string_date_time(value2)
                            
                            if value1_dt == value2_dt:
                                # logging.debug(f"Date-time match at ({row2+1}, {col2+1}): {value1} vs {value2}")
                                continue
                            
                            if re.sub(r'[：【】()（）]', '', str(value1)) != re.sub(r'[：【】()（）]', '', str(value2)):
                                ws.cell(row=row2+1, column=col2+1).fill = fill_pattern_yellow
                                sheet_mismatch_found += 1
                                mismatch_found += 1
                                sheet_report.append({
                                    "row1": row1+1, "col1": col1+1, "val1": value1,
                                    "row2": row2+1, "col2": col2+1, "val2": value2
                                })
                                logging.info(f"Value mismatch at ({row2+1}, {col2+1}): {value1} vs {value2}")
            
            if sheet_report:
                reports.append({
                    "sheet_name": sheet_name,
                    "sheet_report": sheet_report,
                    "mismatch_found": sheet_mismatch_found
                })
                
        
        result = 'X' if mismatch_found > 0 else 'O'
        logging.info(f"Comparison completed. Result: {result} (mismatches: {mismatch_found})")
        return result, wb2_openpyxl, reports
            
    except Exception as e:
        logging.error(f"Error comparing {file1} and {file2}: {e}")
        return None, None, []

def generate_excel_report(all_reports: List[Dict[str, Optional[List[Dict[str, Any]]]]],report_dir: str) -> Optional[str]:
    """
    Generates an Excel report of comparison results with separate sheets for each school ID, containing workbook tables.
    
    Args:
        all_reports: List of dictionaries, each mapping school IDs to lists of file comparison reports.
                     Expected structure: [{school_id: [{workbook_name: [sheet_reports]}]}]
                     where sheet_reports contain 'sheet_name', 'sheet_report', and 'mismatch_found'.
    
    Returns:
        Optional[str]: Path to the generated Excel report file, or None if an error occurs.
    
    Examples:
        >>> reports = [{"50028001": [{"workbook1.xlsx": [{"sheet_name": "Sheet1", "sheet_report": [{"row1": 2, "col1": 3, "val1": "A", "row2": 2, "col2": 3, "val2": "B"}], "mismatch_found": 1}]}]}]
        >>> report_path = generate_excel_report(reports)
        >>> print(report_path)
        '/path/to/20250606_154935_comparison_report.xlsx'
    """
    try:
        # Validate input
        if not isinstance(all_reports, list):
            logging.error(f"Invalid input: all_reports must be a list, got {type(all_reports)}")
            return None

        wb = openpyxl.Workbook()

        # Create summary sheet
        summary_sheet: Worksheet = wb.active
        summary_sheet.title = "Summary"
        summary_headers = ["School ID", "Mismatch Count", "Status", "File Count"]
        for col, header in enumerate(summary_headers, 1):
            cell = summary_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        summary_row = 2

        if not all_reports:
            cell = summary_sheet.cell(row=summary_row, column=1)
            cell.value = "No mismatches found"
            cell.alignment = Alignment(horizontal='center')
            summary_sheet.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
            logging.info("No reports provided, created summary with 'No mismatches found'")
        else:
            for school_report in all_reports:
                if not isinstance(school_report, dict):
                    logging.warning(f"Skipping invalid school report: {school_report}")
                    continue
                
                for school_id, file_reports in school_report.items():
                    if not isinstance(school_id, str) or not isinstance(file_reports, list) :
                        logging.debug(f"Skipping invalid or empty report for school_id: {school_id}")
                        continue
                    
                    # Sanitize sheet name
                    sheet_name = re.sub(r'[\\\/:*?"<>|]', '_', str(school_id))[:31]
                    try:
                        school_sheet: Worksheet = wb.create_sheet(sheet_name)
                    except ValueError:
                        for i in range(1, 100):
                            try:
                                school_sheet = wb.create_sheet(f"{sheet_name[:28]}_{i}")
                                break
                            except ValueError:
                                continue
                        else:
                            logging.warning(f"Could not create sheet for school_id: {school_id}. Skipping...")
                            continue
                    
                    current_row = 1
                    school_mismatch_total = 0
                    
                    for file_report in file_reports:
                        if not isinstance(file_report, dict) or not file_report:
                            logging.debug(f"Skipping invalid file report: {file_report}")
                            continue
                        
                        workbook_name = next(iter(file_report.keys()), None)
                        sheets = file_report.get(workbook_name, [])
                        
                        if not workbook_name or not isinstance(sheets, list):
                            logging.debug(f"Skipping invalid workbook data: {workbook_name}")
                            continue
                        
                        if not sheets:
                            logging.debug(f"Skipping empty workbook data: {workbook_name}")
                            continue
                        
                        # Add workbook title
                        school_sheet.cell(row=current_row, column=1).value = f"Workbook: {workbook_name}"
                        school_sheet.cell(row=current_row, column=1).font = Font(bold=True)
                        current_row += 2  # Leave a blank row after title
                        
                        # Set up table headers
                        headers = ["Sheet Name", "Count", "V1 Row", "V1 Col", "V1 Value", "V2 Row", "V2 Col", "V2 Value"]
                        for col, header in enumerate(headers, 1):
                            cell = school_sheet.cell(row=current_row, column=col)
                            cell.value = header
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center')
                        current_row += 1
                        
                        for sheet_report in sheets:
                            if not isinstance(sheet_report, dict):
                                logging.debug(f"Skipping invalid sheet report: {sheet_report}")
                                continue
                            
                            sheet_name = sheet_report.get("sheet_name", "")
                            mismatches = sheet_report.get("sheet_report", [])
                            mismatch_count = sheet_report.get("mismatch_found", 0)
                            
                            if mismatch_count == 0:
                                school_sheet.cell(row=current_row, column=1).value = sheet_name
                                school_sheet.cell(row=current_row, column=2).value = mismatch_count
                                for col in range(3, 9):
                                    school_sheet.cell(row=current_row, column=col).value = "-" if col != 5 else "No mismatches"
                                current_row += 1
                                continue
                            
                            school_mismatch_total += mismatch_count
                            first_mismatch = True
                            
                            for mismatch in mismatches:
                                if not isinstance(mismatch, dict):
                                    logging.debug(f"Skipping invalid mismatch: {mismatch}")
                                    continue
                                
                                row1 = mismatch.get("row1", "-")
                                col1 = mismatch.get("col1", "-")
                                val1 = mismatch.get("val1", "")
                                row2 = mismatch.get("row2", "-")
                                col2 = mismatch.get("col2", "-")
                                val2 = mismatch.get("val2", "")
                                
                                school_sheet.cell(row=current_row, column=1).value = sheet_name if first_mismatch else ""
                                school_sheet.cell(row=current_row, column=2).value = mismatch_count if first_mismatch else ""
                                school_sheet.cell(row=current_row, column=3).value = row1
                                school_sheet.cell(row=current_row, column=4).value = col1
                                school_sheet.cell(row=current_row, column=5).value = val1
                                school_sheet.cell(row=current_row, column=6).value = row2
                                school_sheet.cell(row=current_row, column=7).value = col2
                                school_sheet.cell(row=current_row, column=8).value = val2
                                first_mismatch = False
                                current_row += 1
                            
                            current_row += 3  # Add three blank rows between workbook tables
                        
                    # Update summary sheet
                    summary_sheet.cell(row=summary_row, column=1).value = school_id
                    summary_sheet.cell(row=summary_row, column=2).value = school_mismatch_total
                    summary_sheet.cell(row=summary_row, column=3).value = "OK" if school_mismatch_total == 0 else "NG"
                    summary_sheet.cell(row=summary_row, column=4).value = len(file_reports)
                    summary_row += 1
        
        # Auto-adjust column widths
        for sheet in wb:
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[column].width = adjusted_width
        
        # Save the Excel report in the reports folder
        os.makedirs(report_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_path = os.path.join(report_dir, f"{timestamp}_comparison_report.xlsx")
        wb.save(report_path)
        logging.info(f"Excel report generated successfully at: {report_path}")
        return report_path
    
    except Exception as e:
        logging.error(f"Error generating Excel report: {e}")
        return None

def generate_report(all_reports: List[Dict[str, Optional[List[Dict[str, Any]]]]],report_dir: str) -> Optional[str]:
    """
    Generates a markdown report of comparison results organized by school ID and workbook.
    
    Args:
        all_reports: List of dictionaries, each mapping school IDs to lists of file comparison reports.
                     Expected structure: [{school_id: [{workbook_name: [sheet_reports]}]}]
                     where sheet_reports contain 'sheet_name', 'sheet_report', and 'mismatch_found'.
    
    Returns:
        Optional[str]: Path to the generated markdown report file, or None if an error occurs.
    
    Examples:
        >>> reports = [{"50028001": [{"workbook1.xlsx": [{"sheet_name": "Sheet1", "sheet_report": [{"row1": 2, "col1": 3, "val1": "A", "row2": 2, "col2": 3, "val2": "B"}], "mismatch_found": 1}]}]}]
        >>> report_path = generate_report(reports)
        >>> print(report_path)
        '/path/to/20250606_154935_comparison_report.md'
    """
    try:
        # Validate input
        if not isinstance(all_reports, list):
            logging.error(f"Invalid input: all_reports must be a list, got {type(all_reports)}")
            return None

        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        report_lines = [
            "# Excel Comparison Report",
            f"**Generated on:** {timestamp}",
            ""
        ]
        
        final_report = [
            "# Final Result Report",
            "",
            "| School ID | Mismatch | Status | File Count |",
            "|-----------|---------|--------|----------|",
        ]
        
        if not all_reports:
            report_lines.append("## No mismatches found")
            report_lines.append("No differences were detected during the comparison.")
            report_path = save_report(report_lines)
            logging.info(f"Report generated at: {report_path}")
            return report_path
        
        for school_report in all_reports:
            if not isinstance(school_report, dict):
                logging.warning(f"Skipping invalid school report: {school_report}")
                continue
            
            for school_id, file_reports in school_report.items():
                if not isinstance(school_id, str)  or not isinstance(file_reports, list) :
                    logging.debug(f"Skipping invalid or empty report for school_id: {school_id}")
                    continue
                
                report_lines.append(f"## School ID: {school_id}")
                report_lines.append("")
                
                school_mismatch_found = 0
                
                for file_report in file_reports:
                    if not isinstance(file_report, dict) or not file_report:
                        logging.debug(f"Skipping invalid file report: {file_report}")
                        continue
                    
                    workbook_name = next(iter(file_report.keys()), None)
                    sheets = file_report.get(workbook_name, [])
                    
                    if not sheets:
                        logging.debug(f"Skipping empty workbook data: {workbook_name}")
                        continue
                    
                    if not workbook_name or not isinstance(sheets, list):
                        logging.debug(f"Skipping invalid workbook data: {workbook_name}")
                        continue
                    
                    report_lines.append(f"### Workbook: {workbook_name}")
                    report_lines.append("")
                    report_lines.append("| Sheet Name | Count | V1 Row | V1 Col | V1 Value | V2 Row | V2 Col | V2 Value |")
                    report_lines.append("|------------|-------|--------|--------|----------|--------|--------|----------|")
                    
                    current_sheet = None
                    
                    for sheet_report in sheets:
                        if not isinstance(sheet_report, dict):
                            logging.debug(f"Skipping invalid sheet report: {sheet_report}")
                            continue
                        
                        sheet_name = sheet_report.get("sheet_name", "")
                        mismatches = sheet_report.get("sheet_report", [])
                        mismatch_count = sheet_report.get("mismatch_found", 0)
                        
                        if mismatch_count == 0:
                            report_lines.append(f"| {sheet_name} | {mismatch_count} | - | - | No mismatches | - | - | - |")
                            continue
                        
                        school_mismatch_found += mismatch_count
                            
                        for mismatch in mismatches:
                            if not isinstance(mismatch, dict):
                                logging.debug(f"Skipping invalid mismatch: {mismatch}")
                                continue
                            
                            row1 = mismatch.get("row1", "-")
                            col1 = mismatch.get("col1", "-")
                            val1 = mismatch.get("val1", "")
                            row2 = mismatch.get("row2", "-")
                            col2 = mismatch.get("col2", "-")
                            val2 = mismatch.get("val2", "")
                            
                            # Escape pipe characters in values to prevent markdown table breakage
                            val1 = str(val1).replace("|", "\\|")
                            val2 = str(val2).replace("|", "\\|")
                            
                            if sheet_name != current_sheet:
                                if current_sheet is not None:
                                    report_lines.append("| | | | | | | | |")
                                current_sheet = sheet_name
                                report_lines.append(
                                    f"| {sheet_name} | {mismatch_count} | {row1} | {col1} | {val1} | {row2} | {col2} | {val2} |"
                                )
                            else:
                                report_lines.append(
                                    f"| | | {row1} | {col1} | {val1} | {row2} | {col2} | {val2} |"
                                )
                    
                    report_lines.append("") 
                
                status = "OK" if school_mismatch_found == 0 else "NG"
                final_report.append(f"| {school_id} | {school_mismatch_found} | {status} | {len(file_reports)}")
            
        report_lines.append("")
        report_lines.extend(final_report)
        
        report_path = save_report(report_lines,report_dir)
        logging.info(f"Report generated at: {report_path}")
        return report_path
    
    except OSError as e:
        logging.error(f"Error generating markdown report: {e}")
        return None

def save_report(report_lines: List[str],report_dir: str) -> str:
    """
    Saves a list of report lines to a markdown file with a timestamped filename.
    
    Args:
        report_lines: List of strings representing the lines of the markdown report.
    
    Returns:
        str: Path to the saved markdown report file.
    
    Raises:
        OSError: If the file cannot be written (e.g., due to permissions or disk issues).
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    os.makedirs(report_dir, exist_ok=True)
    report_path = os.path.join(report_dir, f"{timestamp}_comparison_report.md")
    
    try:
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(report_lines))
        logging.debug(f"Markdown report saved to: {report_path}")
    except OSError as e:
        logging.error(f"Failed to save report to {report_path}: {e}")
        raise
    
    return report_path
    
def process_folder(compare_folder: str,report_dir: str) -> bool:
    """
    Processes subfolders in the specified directory, comparing Excel files in V1 and V2 folders,
    saving results, and generating markdown and Excel reports.
    
    Args:
        compare_folder: Path to the directory containing subfolders with V1 and V2 directories.
    
    Returns:
        bool: True if processing completes successfully, False if an error occurs.
    
    Examples:
        >>> success = process_folder('/path/to/compare_folder')
        >>> print(success)
        True
    """
    try:
        # Validate input
        if not isinstance(compare_folder, str) or not os.path.isdir(compare_folder):
            logging.error(f"Invalid or non-existent folder: {compare_folder}")
            return False

        all_reports: List[Dict[str, List[Dict[str, Any]]]] = []
        subfolders = [f for f in os.listdir(compare_folder) if os.path.isdir(os.path.join(compare_folder, f))]
        
        if not subfolders:
            logging.warning(f"No subfolders found in {compare_folder}")
            return True

        for subfolder in subfolders:
            subfolder_path = os.path.join(compare_folder, subfolder)
            v1_path = os.path.join(subfolder_path, 'V1')
            v2_path = os.path.join(subfolder_path, 'V2')
            result_path = os.path.join(subfolder_path, 'result')
            
            # Check if V1 and V2 folders exist
            if not all(os.path.exists(p) for p in [v1_path, v2_path]):
                logging.debug(f"Skipping {subfolder}: V1 or V2 folder missing")
                continue
            
            # Create result folder if it doesn't exist
            try:
                os.makedirs(result_path, exist_ok=True)
                logging.debug(f"Created result folder: {result_path}")
            except OSError as e:
                logging.error(f"Failed to create result folder {result_path}: {e}")
                continue
            
            # Get Excel files from V1 and V2
            files_v1 = [f for f in os.listdir(v1_path) if f.lower().endswith(('.xlsx', '.xls'))]
            files_v2 = [f for f in os.listdir(v2_path) if f.lower().endswith(('.xlsx', '.xls'))]
            
            school_reports: List[Dict[str, List[Dict]]] = []
            
            for file_name in files_v1:
                base_name = os.path.splitext(file_name)[0]
                matching_files = [f for f in files_v2 if os.path.splitext(f)[0] == base_name]
                
                if not matching_files:
                    logging.debug(f"No matching V2 file found for {file_name} in {subfolder}")
                    continue
                
                file2_name = matching_files[0]
                logging.info(f"Processing files: {file_name} vs {file2_name} in {subfolder}")
                
                file1 = os.path.join(v1_path, file_name)
                file2 = os.path.join(v2_path, file2_name)
                
                try:
                    result, modified_wb, reports = compare_excel_files(file1, file2, subfolder)
                    
                    if result is None or modified_wb is None:
                        logging.warning(f"Skipping save for {file_name} due to comparison error")
                        continue
                    
                    output_path = os.path.join(result_path, f"{result}_{base_name}.xlsx")
                    
                    logging.debug(f"Saving comparison result to: {output_path}")
                    try:
                        modified_wb.save(output_path)
                        logging.info(f"Saved comparison result to: {output_path}")
                    except OSError as e:
                        logging.error(f"Failed to save workbook to {output_path}: {e}")
                        continue
                    
                    school_reports.append({
                        file2_name: reports
                    })
                    logging.debug(f"Added reports for {file2_name} to school_reports")
                    
                except Exception as e:
                    logging.error(f"Error processing {file1} and {file2}: {e}")
                    continue
            
            all_reports.append({
                subfolder: school_reports
            })
            logging.debug(f"Added school reports for {subfolder} to all_reports")
                
        
        if all_reports:
            try:
                md_report_path = generate_report(all_reports,report_dir)
                excel_report_path = generate_excel_report(all_reports,report_dir)
                if md_report_path and excel_report_path:
                    logging.info(f"Reports generated successfully: Markdown at {md_report_path}, Excel at {excel_report_path}")
                elif md_report_path or excel_report_path:
                    logging.warning(f"Partial report generation: Markdown at {md_report_path}, Excel at {excel_report_path}")
                else:
                    logging.error("Failed to generate both markdown and Excel reports")
            except Exception as e:
                logging.error(f"Error generating reports: {e}")
                return False
        else:
            logging.info("No reports generated due to no valid comparisons")
        
        return True
    
    except Exception as e:
        logging.error(f"Error processing folder {compare_folder}: {e}")
        return False
def main() -> None:
    """
    Main entry point for the Excel comparison program.
    Initializes a GUI, prompts for a folder selection, processes the folder, and displays results.
    
    Returns:
        None
    
    Examples:
        >>> main()
        # Opens a folder selection dialog, processes Excel files, and shows completion messages
    """
    
    logger = setup_logging('DEBUG')
    if logger is None:
        raise RuntimeError("Failed to initialize logging")
    
    logging.info('Starting Excel comparison program')
    
    
    try:
        with create_root() as root:
            compare_folder = select_directory("比較するフォルダを選択してください")
            if not compare_folder or not os.path.isdir(compare_folder):
                logging.warning("No valid folder selected")
                show_message("警告", "フォルダが選択されていません。プログラムを終了します。")
                return
            
            report_dir = select_directory("レポートを保存するフォルダを選択してください")
            if not report_dir or not os.path.isdir(report_dir):
                logging.warning("No valid folder selected")
                show_message("警告", "フォルダが選択されていません。プログラムを終了します。")
                return
            
            logging.info(f"Starting comparison process for folder: {compare_folder}")
            show_message("情報", "比較処理を開始しました。")
            
            if process_folder(compare_folder,report_dir):
                logging.info("Comparison process completed successfully")
                show_message("情報", "比較処理が正常に完了しました。")
            else:
                logging.error("Comparison process failed")
                show_message("エラー", "比較処理に失敗しました。")
    
    except Exception as e:
        logging.error(f"An error occurred in main: {e}")
        show_message("エラー", f"エラーが発生しました: {str(e)}")
    
    finally:
        logging.info("Program finished")

if __name__ == "__main__":
    main()